"""JSON 与 Excel 审核报告导出。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from common import ACTIVE_AUDIT_CONFIG, APP_VERSION, MANUAL, PASS, RISK, SKIP, AuditError


def excel_safe(value: Any) -> Any:
    """防止来自模型或转写的文本被电子表格当作公式执行。"""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value

def save_json(result: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


def evidence_text(item: dict[str, Any]) -> str:
    return "；".join(
        f"{entry.get('start', '')}-{entry.get('end', '')} {entry.get('detail', '')}"
        for entry in item.get("evidence", [])
    )


def save_excel(results: list[dict[str, Any]], path: Path, script_pdf: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise AuditError("缺少 openpyxl，请先安装 requirements.txt。") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "审核汇总"
    summary.append(
        [
            "总体结论",
            "；".join(
                f"{result['video']['file_name']}：{result['overall_status']}"
                for result in results
            ),
        ]
    )
    summary.append([])
    summary.append(
        [
            "视频",
            "规则编号",
            "类别",
            "审核项",
            "审核方式",
            "最终结论",
            "来源结论",
            "置信度",
            "原因",
            "证据时间段",
        ]
    )
    for result in results:
        for item in result["aggregate"]:
            source_statuses = "；".join(
                f"{source}:{status}"
                for source, status in item.get("source_statuses", {}).items()
            )
            summary.append(
                [
                    result["video"]["file_name"],
                    item["rule_id"],
                    item["category"],
                    item["name"],
                    item["method"],
                    item["status"],
                    source_statuses,
                    item["confidence"],
                    item["reason"],
                    evidence_text(item),
                ]
            )

    visual_sheet = workbook.create_sheet("视频画面审核")
    visual_sheet.append(
        ["视频", "模型", "规则编号", "审核项", "结论", "置信度", "原因", "证据"]
    )
    for result in results:
        for item in result["visual"]["rules"]:
            visual_sheet.append(
                [
                    result["video"]["file_name"],
                    result["visual"]["model"],
                    item["rule_id"],
                    item["name"],
                    item["status"],
                    item["confidence"],
                    item["reason"],
                    evidence_text(item),
                ]
            )

    audio_sheet = workbook.create_sheet("音频审核")
    audio_sheet.append(
        ["视频", "ASR模型", "分析模型", "规则编号", "审核项", "结论", "置信度", "原因", "证据"]
    )
    for result in results:
        for item in result["audio"]["rules"]:
            audio_sheet.append(
                [
                    result["video"]["file_name"],
                    result["audio"]["asr_model"],
                    result["audio"]["analysis_model"],
                    item["rule_id"],
                    item["name"],
                    item["status"],
                    item["confidence"],
                    item["reason"],
                    evidence_text(item),
                ]
            )

    script_enabled = bool(
        ACTIVE_AUDIT_CONFIG.get("script_comparison", {}).get("enabled", False)
    )
    script_rows = [
        (result, item)
        for result in results
        for item in result["audio"].get("script_checks", [])
    ] if script_enabled else []
    create_empty_script_sheet = bool(
        ACTIVE_AUDIT_CONFIG.get("output", {}).get(
            "create_empty_script_sheet", False
        )
    )
    if script_rows or create_empty_script_sheet:
        script_sheet = workbook.create_sheet("话术比对")
        script_sheet.append(
            ["视频", "标准事项", "比对结论", "实际语音摘要", "开始", "结束", "原因"]
        )
        for result, item in script_rows:
            script_sheet.append(
                [
                    result["video"]["file_name"], item["item"], item["status"],
                    item["heard_text"], item["start"], item["end"], item["reason"],
                ]
            )

    include_transcript = bool(
        ACTIVE_AUDIT_CONFIG.get("output", {}).get("include_transcript", True)
    )
    if include_transcript:
        raw_transcript_sheet = workbook.create_sheet("原始ASR转写")
        raw_transcript_sheet.append(
            ["视频", "片段编号", "开始", "结束", "原始转写"]
        )
        for result in results:
            for item in result["audio"].get("asr_sentences", []):
                raw_transcript_sheet.append(
                    [result["video"]["file_name"], item.get("segment_id", ""),
                     item.get("start", ""), item.get("end", ""), item.get("text", "")]
                )

        transcript_sheet = workbook.create_sheet("对话转写")
        transcript_sheet.append(
            ["视频", "片段编号", "说话人", "开始", "结束", "转写"]
        )
        for result in results:
            for item in result["audio"].get("dialogue", []):
                transcript_sheet.append(
                    [result["video"]["file_name"], item.get("segment_id", ""),
                     item.get("speaker", "unknown"), item.get("start", ""),
                     item.get("end", ""), item.get("text", "")]
                )

    routing_sheet = workbook.create_sheet("规则分流与人工项")
    routing_sheet.append(["视频", "分流状态", "类型", "内容"])
    for result in results:
        routing = result.get("rule_routing", {})
        if not isinstance(routing, dict):
            continue
        rows = [
            ("视觉要求", routing.get("visual_requirements", "")),
            ("音频要求", routing.get("audio_requirements", "")),
            ("综合要求", routing.get("combined_requirements", "")),
            ("人工/系统要求", routing.get("manual_or_system_requirements", "")),
        ]
        rows.extend(("无法分类", value) for value in routing.get("uncertain_items", []))
        for category, content in rows:
            if str(content).strip():
                routing_sheet.append(
                    [result["video"]["file_name"], routing.get("routing_status", ""),
                     category, str(content)]
                )

    review_sheet = workbook.create_sheet("待复核清单")
    review_sheet.append(
        ["视频", "规则编号", "类别", "审核项", "最终结论", "原因", "证据时间段", "建议操作"]
    )
    review_rows: list[tuple[int, list[Any]]] = []
    priority = {RISK: 1, MANUAL: 2, SKIP: 3}
    for result in results:
        for item in result["aggregate"]:
            if item["status"] not in priority:
                continue
            suggestion = (
                "立即复核"
                if item["status"] == RISK
                else "人工确认"
                if item["status"] == MANUAL
                else "需系统数据"
            )
            review_rows.append(
                (
                    priority[item["status"]],
                    [
                        result["video"]["file_name"],
                        item["rule_id"],
                        item["category"],
                        item["name"],
                        item["status"],
                        item["reason"],
                        evidence_text(item),
                        suggestion,
                    ],
                )
            )
    for _, row in sorted(review_rows, key=lambda entry: entry[0]):
        review_sheet.append(row)

    info_sheet = workbook.create_sheet("运行信息")
    info_sheet.append(["视频", "项目", "值"])
    for result in results:
        video_name = result["video"]["file_name"]
        transcript_integrity = result["audio"].get("transcript_integrity", {})
        info_rows = [
            ("Demo版本", APP_VERSION),
            ("生成时间", result["finished_at"]),
            ("业务场景", result["scenario"]["name"]),
            ("视频模型", result["models"]["visual"]),
            ("ASR模型", result["models"]["asr"]),
            ("处理方式", result["parameters"]["processing_mode"]),
            ("审核范围", "完整视频全局统一审核"),
            ("视频抽帧率", result["parameters"]["fps"]),
            ("音频时长秒", result["audio"]["metrics"].get("duration_seconds", "")),
            ("音频RMS", result["audio"]["metrics"].get("rms_normalized", "")),
            ("静音占比", result["audio"]["metrics"].get("silence_ratio", "")),
            ("削波占比", result["audio"]["metrics"].get("clipping_ratio", "")),
            ("ASR原始片段数", transcript_integrity.get("asr_segment_count", "")),
            ("对话转写片段数", transcript_integrity.get("dialogue_segment_count", "")),
            ("转写文本覆盖率", transcript_integrity.get("text_coverage", "")),
            ("说话人标签覆盖率", transcript_integrity.get("speaker_label_coverage", "")),
            ("unknown说话人数", transcript_integrity.get("unknown_speaker_count", "")),
            ("标准话术文件", script_pdf.name),
            ("数据说明", "报告保留审核所需真实信息；不会写入 API Key。请按敏感资料保管。"),
        ]
        for key, value in info_rows:
            info_sheet.append([video_name, key, value])
            if "覆盖率" in key and isinstance(value, (int, float)):
                info_sheet.cell(row=info_sheet.max_row, column=3).number_format = "0.0%"

    status_fills = {
        PASS: PatternFill("solid", fgColor="C6EFCE"),
        RISK: PatternFill("solid", fgColor="FFC7CE"),
        MANUAL: PatternFill("solid", fgColor="FFEB9C"),
        SKIP: PatternFill("solid", fgColor="D9EAF7"),
        "部分审核通过": PatternFill("solid", fgColor="D9EAD3"),
    }
    for sheet in workbook.worksheets:
        header_row = 3 if sheet.title == "审核汇总" else 1
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = excel_safe(cell.value)
        sheet.freeze_panes = f"A{header_row + 1}"
        if sheet.max_row >= header_row:
            sheet.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            )
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if sheet.title == "审核汇总":
            summary["A1"].font = Font(bold=True, color="FFFFFF")
            summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
            summary["B1"].font = Font(bold=True)
            overall_values = {result["overall_status"] for result in results}
            if RISK in overall_values:
                summary["B1"].fill = status_fills[RISK]
            elif MANUAL in overall_values:
                summary["B1"].fill = status_fills[MANUAL]
            else:
                summary["B1"].fill = status_fills["部分审核通过"]
        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.value in status_fills:
                    cell.fill = status_fills[cell.value]
        for column_index, column_cells in enumerate(sheet.columns, start=1):
            values = [str(cell.value or "") for cell in list(column_cells)[:200]]
            width = min(60, max(10, max((len(value) for value in values), default=10) + 2))
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
