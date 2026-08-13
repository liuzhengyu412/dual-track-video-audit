#!/usr/bin/env python3
"""配置驱动的视频音频双轨审核引擎。"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from common import (
    ACTIVE_AUDIT_CONFIG,
    APP_VERSION,
    AUDIO_RULE_IDS,
    MANUAL,
    PASS,
    RISK,
    RULES,
    RULE_BY_ID,
    SCENARIOS,
    SKIP,
    SYSTEM_SKIP_REASONS,
    VISUAL_RULE_IDS,
    AuditError,
    Rule,
    bypass_broken_proxy_for_dashscope,
    load_dotenv,
    ms_to_timestamp,
    safe_error_text,
    sha256_file,
)
from media import prepare_audio as prepare_media_audio
from models import (
    call_qwen,
    decumulate_asr_sentences,
    extract_qwen_text,
    parse_json_object,
    response_to_dict,
    transcribe_audio_chunks,
)
from reporting import save_excel, save_json
from results import (
    build_dialogue_from_asr,
    combine_audit_results,
    correction_prompt,
    make_failure_source,
    normalize_rule_results,
    normalize_script_checks,
    normalize_timestamp_for_duration,
    normalize_timestamps_in_text,
    overall_status,
    repair_rule_payload_timestamps,
    routing_review_items,
    sanitize_asr_sentences,
    timestamp_seconds,
    validate_rule_payload,
    validate_speaker_labels,
)


PROJECT_DIR = Path(__file__).resolve().parent
WORK_DIR = PROJECT_DIR / "work"
DEFAULT_OUTPUT_DIR = PROJECT_DIR / "results"
DEFAULT_CONFIG_PATH = PROJECT_DIR / "rules" / "audit_config.json"
DEFAULT_SETTINGS_PATH = PROJECT_DIR / "settings.json"
DEFAULT_VISUAL_MODEL = "qwen3-vl-32b-instruct"
DEFAULT_ASR_MODEL = "fun-asr-realtime"
DEFAULT_API_BASE = "https://dashscope.aliyuncs.com/api/v1"

DEFAULT_SCENARIO = ""


def load_audit_config(config_path: Path) -> dict[str, Any]:
    """加载可由使用者维护的规则、场景和提示词配置。"""
    if not config_path.is_file():
        raise AuditError(f"找不到审核配置文件：{config_path}")
    try:
        payload = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise AuditError(f"无法读取审核配置文件：{config_path}") from exc
    if not isinstance(payload, dict):
        raise AuditError("审核配置必须是JSON对象。")
    if not DEFAULT_SETTINGS_PATH.is_file():
        raise AuditError(f"找不到项目设置文件：{DEFAULT_SETTINGS_PATH}")
    try:
        settings = json.loads(DEFAULT_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise AuditError(f"无法读取项目设置文件：{DEFAULT_SETTINGS_PATH}") from exc
    if not isinstance(settings, dict):
        raise AuditError("settings.json 必须是 JSON 对象。")
    script_comparison = settings.get("script_comparison", {})
    if not isinstance(script_comparison, dict):
        raise AuditError("settings.json 中的 script_comparison 必须是对象。")
    mode = str(settings.get("mode", "simple")).strip().lower()
    if mode not in {"simple", "advanced"}:
        raise AuditError("mode 只能为 simple 或 advanced。")
    routing_failure_strategy = str(
        settings.get("routing_failure_strategy", "fallback_both")
    ).strip().lower()
    if routing_failure_strategy not in {"fallback_both", "stop"}:
        raise AuditError(
            "routing_failure_strategy 只能为 fallback_both 或 stop。"
        )
    limits = settings.get("limits", {})
    if not isinstance(limits, dict):
        raise AuditError("settings.json 中的 limits 必须是对象。")
    try:
        max_transcript_characters = int(
            limits.get("max_transcript_characters", 60000)
        )
    except (TypeError, ValueError) as exc:
        raise AuditError("limits.max_transcript_characters 必须是整数。") from exc
    if max_transcript_characters < 1000:
        raise AuditError("limits.max_transcript_characters 不得小于 1000。")
    output_settings = settings.get("output", {})
    if not isinstance(output_settings, dict):
        raise AuditError("settings.json 中的 output 必须是对象。")
    speaker_roles = settings.get("speaker_roles", ["staff", "customer", "unknown"])
    if (
        not isinstance(speaker_roles, list)
        or not speaker_roles
        or any(not isinstance(role, str) or not role.strip() for role in speaker_roles)
    ):
        raise AuditError("settings.json 中的 speaker_roles 必须是非空字符串数组。")
    speaker_roles = list(dict.fromkeys(role.strip().lower() for role in speaker_roles))
    if "unknown" not in speaker_roles:
        speaker_roles.append("unknown")
    processing = settings.get("processing", {})
    if not isinstance(processing, dict):
        raise AuditError("settings.json 中的 processing 必须是对象。")
    try:
        processing_values = {
            "fps": float(processing.get("fps", 2.0)),
            "audio_chunk_seconds": int(processing.get("audio_chunk_seconds", 180)),
            "confidence_threshold": float(processing.get("confidence_threshold", 0.70)),
            "max_video_size_mb": int(processing.get("max_video_size_mb", 100)),
            "keep_work_files": bool(processing.get("keep_work_files", False)),
        }
    except (TypeError, ValueError) as exc:
        raise AuditError("settings.json 中的 processing 参数类型无效。") from exc
    if not 0.1 <= processing_values["fps"] <= 10:
        raise AuditError("processing.fps 必须在 0.1 到 10 之间。")
    if not 30 <= processing_values["audio_chunk_seconds"] <= 240:
        raise AuditError("processing.audio_chunk_seconds 必须在 30 到 240 之间。")
    if not 0 <= processing_values["confidence_threshold"] <= 1:
        raise AuditError("processing.confidence_threshold 必须在 0 到 1 之间。")
    if processing_values["max_video_size_mb"] < 1:
        raise AuditError("processing.max_video_size_mb 必须大于 0。")
    base_dir = config_path.parent
    review_brief_path: Path | None = None
    if mode == "simple":
        review_brief = str(payload.get("review_brief", "")).strip()
        if not review_brief:
            raise AuditError("简易模式需要填写 review_brief（审核要求文件路径）。")
        review_brief_path = (base_dir / review_brief).resolve()
        if not review_brief_path.is_file():
            raise AuditError(f"找不到审核要求文件：{review_brief_path}")
        review_brief_text = review_brief_path.read_text(encoding="utf-8").strip()
        if not review_brief_text:
            raise AuditError(f"审核要求文件为空：{review_brief_path}")
        template_incomplete = "<!-- REVIEW_RULES_REQUIRED -->" in review_brief_text
        raw_rules = [
            {"id": "visual_review", "category": "视觉审核", "name": "画面是否符合审核要求", "method": "视觉模型", "tracks": ["visual"]},
            {"id": "audio_review", "category": "音频审核", "name": "语音是否符合审核要求", "method": "音频模型", "tracks": ["audio"]},
            {"id": "cross_modal_review", "category": "综合审核", "name": "跨轨证据是否存在冲突或风险", "method": "视觉与音频模型", "tracks": ["visual", "audio"]},
        ]
        payload["global_instruction"] = review_brief_text
    else:
        template_incomplete = False
        raw_rules = payload.get("rules")
        if not isinstance(raw_rules, list) or not raw_rules:
            raise AuditError("高级模式至少需要一条 rules。")

    rules: list[Rule] = []
    tracks_by_id: dict[str, tuple[str, ...]] = {}
    skip_reasons: dict[str, str] = {}
    instructions: dict[str, str] = {}
    seen: set[str] = set()
    for index, item in enumerate(raw_rules, start=1):
        if not isinstance(item, dict):
            raise AuditError(f"rules[{index}] 必须是对象。")
        rule_id = str(item.get("id", "")).strip()
        name = str(item.get("name", "")).strip()
        category = str(item.get("category", "未分类")).strip() or "未分类"
        method = str(item.get("method", "模型审核")).strip() or "模型审核"
        tracks = item.get("tracks", [])
        if not rule_id or not name or rule_id in seen:
            raise AuditError(f"rules[{index}] 的 id 必须唯一，且 id、name 不可为空。")
        if not isinstance(tracks, list) or not tracks:
            raise AuditError(f"规则 {rule_id} 必须至少指定一个 tracks。")
        normalized_tracks = tuple(str(track).strip() for track in tracks)
        allowed_tracks = {"visual", "audio", "system"}
        if any(track not in allowed_tracks for track in normalized_tracks):
            raise AuditError(f"规则 {rule_id} 的 tracks 只能为 visual、audio 或 system。")
        seen.add(rule_id)
        rules.append(Rule(rule_id, category, name, method))
        tracks_by_id[rule_id] = normalized_tracks
        instruction = str(item.get("instruction", "")).strip()
        if instruction:
            instructions[rule_id] = instruction
        if "system" in normalized_tracks:
            skip_reasons[rule_id] = str(
                item.get("skip_reason", "需要外部系统数据，当前未接入。")
            ).strip()

    raw_scenarios = payload.get("scenarios", {})
    if not isinstance(raw_scenarios, dict) or not raw_scenarios:
        raise AuditError("审核配置至少需要一个 scenarios 场景。")
    scenarios: dict[str, dict[str, Any]] = {}
    for scenario_id, item in raw_scenarios.items():
        if not isinstance(item, dict):
            raise AuditError(f"场景 {scenario_id} 必须是对象。")
        name = str(item.get("name", scenario_id)).strip() or str(scenario_id)
        script = str(item.get("script", "")).strip()
        if not script and bool(script_comparison.get("enabled", False)):
            raise AuditError(f"场景 {scenario_id} 必须指定 script 文件。")
        scenarios[str(scenario_id)] = {
            "name": name,
            "script": script,
        }
    prompts = payload.get("prompts", {})
    if not isinstance(prompts, dict):
        raise AuditError("prompts 必须是对象。")
    return {
        "path": config_path.resolve(),
        "rules": tuple(rules),
        "tracks_by_id": tracks_by_id,
        "skip_reasons": skip_reasons,
        "instructions": instructions,
        "scenarios": scenarios,
        "prompts": {str(key): str(value) for key, value in prompts.items()},
        "local_validators": payload.get("local_validators", {}),
        "mode": mode,
        "review_brief_path": str(review_brief_path) if review_brief_path else "",
        "template_incomplete": template_incomplete,
        "default_scenario": str(settings.get("default_scenario", "")).strip(),
        "rule_routing": str(settings.get("rule_routing", "ai")).strip().lower(),
        "routing_failure_strategy": routing_failure_strategy,
        "script_comparison": script_comparison,
        "limits": {"max_transcript_characters": max_transcript_characters},
        "output": output_settings,
        "speaker_roles": speaker_roles,
        "processing": processing_values,
        "settings_path": str(DEFAULT_SETTINGS_PATH.resolve()),
    }


def configure_audit_config(config_path: Path) -> dict[str, Any]:
    """将配置应用到本次进程；CLI、自检和报表共享同一份规则。"""
    config = load_audit_config(config_path)
    global DEFAULT_SCENARIO
    RULES[:] = config["rules"]
    RULE_BY_ID.clear()
    RULE_BY_ID.update({rule.rule_id: rule for rule in RULES})
    VISUAL_RULE_IDS[:] = [
        rule.rule_id for rule in RULES if "visual" in config["tracks_by_id"][rule.rule_id]
    ]
    AUDIO_RULE_IDS[:] = [
        rule.rule_id for rule in RULES if "audio" in config["tracks_by_id"][rule.rule_id]
    ]
    SYSTEM_SKIP_REASONS.clear()
    SYSTEM_SKIP_REASONS.update(config["skip_reasons"])
    SCENARIOS.clear()
    SCENARIOS.update(config["scenarios"])
    DEFAULT_SCENARIO = config["default_scenario"]
    ACTIVE_AUDIT_CONFIG.clear()
    ACTIVE_AUDIT_CONFIG.update(config)
    return config


def local_validator_config(name: str) -> dict[str, Any] | None:
    validators = ACTIVE_AUDIT_CONFIG.get("local_validators", {})
    if not isinstance(validators, dict):
        raise AuditError("local_validators 必须是对象。")
    value = validators.get(name)
    return value if isinstance(value, dict) and value.get("enabled", False) else None


def script_comparison_enabled() -> bool:
    value = ACTIVE_AUDIT_CONFIG.get("script_comparison", {})
    if not isinstance(value, dict):
        raise AuditError("settings.json 中的 script_comparison 必须是对象。")
    return bool(value.get("enabled", False))

def require_api_key() -> str:
    api_key = os.getenv("DASHSCOPE_API_KEY", "").strip()
    if not api_key:
        raise AuditError(
            "未检测到模型API Key。请打开当前目录的.env，"
            "只在 DASHSCOPE_API_KEY= 后填写百炼 API Key。"
        )
    return api_key


def normalize_api_base(value: str) -> str:
    base = value.strip().rstrip("/") or DEFAULT_API_BASE
    if not base.endswith("/api/v1"):
        raise AuditError(
            "MODEL_API_BASE 应以 /api/v1 结尾，例如 "
            "https://dashscope.aliyuncs.com/api/v1"
        )
    return base


def extract_script_text(script_path: Path, scenario: str) -> str:
    if scenario not in SCENARIOS:
        raise AuditError(f"未知话术场景：{scenario}")
    if not script_path.is_file():
        raise AuditError(f"找不到场景话术文件：{script_path}")
    if script_path.suffix.lower() not in {".md", ".txt"}:
        raise AuditError("当前通用版本的话术文件仅支持 UTF-8 编码的 .md 或 .txt。")
    text = script_path.read_text(encoding="utf-8").strip()
    if not text:
        raise AuditError(f"场景话术文件为空：{script_path}")
    return text


def render_prompt_template(template_name: str, values: dict[str, str]) -> str:
    config = ACTIVE_AUDIT_CONFIG
    prompt_files = config.get("prompts", {})
    relative_path = prompt_files.get(template_name)
    if not relative_path:
        raise AuditError(f"审核配置缺少 prompts.{template_name}。")
    template_path = Path(config["path"]).parent / relative_path
    if not template_path.is_file():
        raise AuditError(f"找不到提示词模板：{template_path}")
    template = template_path.read_text(encoding="utf-8")
    for key, value in values.items():
        template = template.replace("{{" + key + "}}", value)
    unresolved = re.findall(r"\{\{[A-Za-z_][A-Za-z0-9_]*\}\}", template)
    if unresolved:
        raise AuditError("提示词模板包含未填充变量：" + ", ".join(sorted(set(unresolved))))
    return template.strip()


def rule_instructions(rule_ids: tuple[str, ...]) -> str:
    config = ACTIVE_AUDIT_CONFIG
    instructions = config.get("instructions", {})
    return "\n".join(
        f"- {rule_id}｜{instructions.get(rule_id, RULE_BY_ID[rule_id].name)}"
        for rule_id in rule_ids
    )


def global_instruction(track: str) -> str:
    config = ACTIVE_AUDIT_CONFIG
    routing = config.get("rule_routing_result", {})
    if isinstance(routing, dict):
        key = {"visual": "visual_requirements", "audio": "audio_requirements"}.get(track)
        if key and str(routing.get(key, "")).strip():
            combined = str(routing.get("combined_requirements", "")).strip()
            content = str(routing[key]).strip()
            return content + ("\n\n同时适用的综合要求：\n" + combined if combined else "")
    return str(config.get("global_instruction", "")).strip() or "无额外通用业务口径。"


def build_rule_routing_prompt() -> str:
    config = ACTIVE_AUDIT_CONFIG
    return render_prompt_template(
        "routing",
        {"review_brief": str(config.get("global_instruction", "")).strip()},
    )


def prepare_audio(
    video_path: Path,
    work_dir: Path,
    chunk_seconds: int,
) -> dict[str, Any]:
    return prepare_media_audio(
        video_path, work_dir, chunk_seconds,
        error_type=AuditError, safe_error_text=safe_error_text,
    )


def build_visual_prompt() -> str:
    rule_lines = "\n".join(
        f"- {rule_id}｜{RULE_BY_ID[rule_id].name}" for rule_id in VISUAL_RULE_IDS
    )
    if not ACTIVE_AUDIT_CONFIG:
        raise AuditError("未加载审核配置，无法构建视觉提示词。")
    return render_prompt_template(
        "visual",
        {
            "rule_lines": rule_lines,
            "global_instruction": global_instruction("visual"),
            "rule_instructions": rule_instructions(VISUAL_RULE_IDS),
            "rule_ids": ", ".join(VISUAL_RULE_IDS),
        },
    )


def transcript_for_prompt(sentences: list[dict[str, Any]]) -> str:
    lines = []
    for index, sentence in enumerate(sentences, start=1):
        segment_id = str(sentence.get("segment_id", "")).strip() or f"SEG-{index:03d}"
        lines.append(
            f"[{segment_id}][{sentence.get('start', '')}-{sentence.get('end', '')}] "
            f"{sentence.get('text', '')}"
        )
    return "\n".join(lines)


def build_audio_review_prompt(
    *,
    scenario_name: str,
    script_text: str,
    sentences: list[dict[str, Any]],
    audio_metrics: dict[str, Any],
) -> str:
    rule_lines = "\n".join(
        f"- {rule_id}｜{RULE_BY_ID[rule_id].name}" for rule_id in AUDIO_RULE_IDS
    )
    transcript_text = transcript_for_prompt(sentences)
    if not ACTIVE_AUDIT_CONFIG:
        raise AuditError("未加载审核配置，无法构建音频提示词。")
    max_characters = int(
        ACTIVE_AUDIT_CONFIG.get("limits", {}).get(
            "max_transcript_characters", 60000
        )
    )
    if len(transcript_text) > max_characters:
        raise AuditError(
            f"ASR 转写共 {len(transcript_text)} 个字符，超过配置上限 "
            f"{max_characters}；为避免截断导致误判，已转人工复核。"
        )
    return render_prompt_template(
        "audio",
        {
            "scenario_name": scenario_name,
            "rule_lines": rule_lines,
            "global_instruction": global_instruction("audio"),
            "rule_instructions": rule_instructions(AUDIO_RULE_IDS),
            "audio_metrics": json.dumps(audio_metrics, ensure_ascii=False, indent=2),
            "script_text": script_text if script_comparison_enabled() else "话术比对功能未启用；不要根据话术判断，也不要返回 script_checks。",
            "script_comparison_instruction": (
                "已启用：逐项比对下方场景话术，并返回 script_checks。"
                if script_comparison_enabled()
                else "未启用：不要做话术比对，script_checks 必须返回空数组。"
            ),
            "transcript_text": transcript_text,
            "speaker_roles": "|".join(ACTIVE_AUDIT_CONFIG.get("speaker_roles", [])),
            "rule_ids": ", ".join(AUDIO_RULE_IDS),
        },
    )


def audit_video(
    video_path: Path,
    *,
    script_text: str,
    scenario: str,
    api_key: str,
    api_base: str,
    visual_model: str,
    asr_model: str,
    fps: float,
    audio_chunk_seconds: int,
    confidence_threshold: float,
    keep_work_files: bool,
) -> dict[str, Any]:
    started_at = datetime.now().astimezone()
    audio_analysis_model = str(
        globals().get("AUDIO_ANALYSIS_MODEL", visual_model) or visual_model
    )
    run_dir = WORK_DIR / f"{video_path.stem}-{uuid.uuid4().hex[:8]}"
    run_dir.mkdir(parents=True, exist_ok=False)
    print(f"[{video_path.name}] 正在准备完整音频……", flush=True)

    try:
        prepared = prepare_audio(video_path, run_dir, audio_chunk_seconds)
        duration_seconds = float(prepared["metrics"].get("duration_seconds", 0.0) or 0.0)

        print(
            f"[{video_path.name}] {visual_model}正在审核视频画面……",
            flush=True,
        )
        try:
            visual_prompt = build_visual_prompt()
            visual_output = call_qwen(
                prompt=visual_prompt,
                api_key=api_key,
                model=visual_model,
                api_base=api_base,
                video_path=video_path,
                fps=fps,
            )
            repair_rule_payload_timestamps(
                visual_output, duration_seconds=duration_seconds
            )
            visual_errors = validate_rule_payload(
                visual_output,
                expected_rule_ids=VISUAL_RULE_IDS,
                duration_seconds=duration_seconds,
            )
            if visual_errors:
                print(
                    f"[{video_path.name}] 视觉结果校验失败，正在请求模型纠正一次……",
                    flush=True,
                )
                visual_output = call_qwen(
                    prompt=correction_prompt(
                        visual_prompt,
                        errors=visual_errors,
                        duration_seconds=duration_seconds,
                    ),
                    api_key=api_key,
                    model=visual_model,
                    api_base=api_base,
                    video_path=video_path,
                    fps=fps,
                )
                repair_rule_payload_timestamps(
                    visual_output, duration_seconds=duration_seconds
                )
            visual_source = normalize_rule_results(
                visual_output,
                expected_rule_ids=VISUAL_RULE_IDS,
                source="视频画面",
                confidence_threshold=confidence_threshold,
                duration_seconds=duration_seconds,
            )
        except Exception as exc:
            visual_source = make_failure_source(
                VISUAL_RULE_IDS,
                source="视频画面",
                reason=str(exc),
            )

        asr_result: dict[str, Any]
        audio_output: dict[str, Any] = {}
        script_checks: list[dict[str, str]] = []
        dialogue: list[dict[str, str]] = []
        transcript_integrity: dict[str, Any] = {}
        asr_error: Exception | None = None
        try:
            asr_result = transcribe_audio_chunks(
                chunks=prepared["chunks"],
                api_key=api_key,
                model=asr_model,
                api_base=api_base,
            )
        except Exception as exc:
            asr_error = exc
            asr_result = {
                "model": asr_model,
                "request_ids": [],
                "sentences": [],
                "text": "",
                "chunk_count": len(prepared["chunks"]),
                "error": safe_error_text(str(exc)),
            }

        if asr_error is None:
            try:
                print(
                    f"[{video_path.name}] {audio_analysis_model}正在根据"
                    f"{asr_result.get('model', asr_model)}转写审核语音规则……",
                    flush=True,
                )
                audio_prompt = build_audio_review_prompt(
                    scenario_name=SCENARIOS[scenario]["name"],
                    script_text=script_text,
                    sentences=asr_result["sentences"],
                    audio_metrics=prepared["metrics"],
                )
                audio_output = call_qwen(
                    prompt=audio_prompt,
                    api_key=api_key,
                    model=audio_analysis_model,
                    api_base=api_base,
                )
                repair_rule_payload_timestamps(
                    audio_output, duration_seconds=duration_seconds
                )
                audio_errors = validate_rule_payload(
                    audio_output,
                    expected_rule_ids=AUDIO_RULE_IDS,
                    duration_seconds=duration_seconds,
                )
                audio_errors.extend(
                    validate_speaker_labels(audio_output, asr_result["sentences"])
                )
                if audio_errors:
                    print(
                        f"[{video_path.name}] 音频规则结果校验失败，正在请求模型纠正一次……",
                        flush=True,
                    )
                    audio_output = call_qwen(
                        prompt=correction_prompt(
                            audio_prompt,
                            errors=audio_errors,
                            duration_seconds=duration_seconds,
                        ),
                        api_key=api_key,
                        model=audio_analysis_model,
                        api_base=api_base,
                    )
                    repair_rule_payload_timestamps(
                        audio_output, duration_seconds=duration_seconds
                    )
                audio_source = normalize_rule_results(
                    audio_output,
                    expected_rule_ids=AUDIO_RULE_IDS,
                    source="音频转写",
                    confidence_threshold=confidence_threshold,
                    duration_seconds=duration_seconds,
                )
                script_checks = normalize_script_checks(
                    audio_output.get("script_checks", [])
                )
                dialogue, transcript_integrity = build_dialogue_from_asr(
                    asr_result.get("sentences", []),
                    audio_output.get("speaker_labels", []),
                )
            except Exception as exc:
                audio_source = make_failure_source(
                    AUDIO_RULE_IDS,
                    source="音频转写",
                    reason=f"转写成功，但音频规则分析失败：{exc}",
                )
        else:
            audio_source = make_failure_source(
                AUDIO_RULE_IDS,
                source="音频转写",
                reason=str(asr_error),
            )

        if not transcript_integrity:
            dialogue, transcript_integrity = build_dialogue_from_asr(
                asr_result.get("sentences", []), []
            )

        effective_asr_model = str(asr_result.get("model", asr_model) or asr_model)

        aggregate = combine_audit_results(
            visual_source=visual_source,
            audio_source=audio_source,
            script_checks=script_checks,
            audio_metrics=prepared["metrics"],
        )
        aggregate.extend(
            routing_review_items(ACTIVE_AUDIT_CONFIG.get("rule_routing_result"))
        )
        finished_at = datetime.now().astimezone()
        result = {
            "video": {
                "file_name": video_path.name,
                "sha256": sha256_file(video_path),
                "size_bytes": video_path.stat().st_size,
            },
            "scenario": {"id": scenario, "name": SCENARIOS[scenario]["name"]},
            "models": {
                "visual": visual_model,
                "asr": effective_asr_model,
                "audio_rule_analysis": audio_analysis_model,
            },
            "parameters": {
                "fps": fps,
                "confidence_threshold": confidence_threshold,
                "processing_mode": "完整视频统一审核；音频内部传输后恢复连续时间轴并合并",
            },
            "started_at": started_at.isoformat(timespec="seconds"),
            "finished_at": finished_at.isoformat(timespec="seconds"),
            "overall_status": overall_status(aggregate),
            "aggregate": aggregate,
            "visual": {
                "model": visual_model,
                "rules": visual_source["rules"],
                "error": visual_source.get("error", ""),
            },
            "audio": {
                "asr_model": effective_asr_model,
                "analysis_model": audio_analysis_model,
                "rules": audio_source["rules"],
                "error": audio_source.get("error", asr_result.get("error", "")),
                "metrics": prepared["metrics"],
                "script_checks": script_checks,
                "asr_sentences": sanitize_asr_sentences(asr_result.get("sentences", [])),
                "dialogue": dialogue,
                "transcript_integrity": transcript_integrity,
            },
        }
        if keep_work_files:
            result["work_files"] = {"directory": str(run_dir.resolve())}
        return result
    finally:
        if not keep_work_files:
            shutil.rmtree(run_dir, ignore_errors=True)


def next_version(output_dir: Path, video_stem: str) -> int:
    versions: list[int] = []
    patterns = (
        f"视频审核汇总_{video_stem}_v*.xlsx",
        f"{video_stem}_分离审核结果_v*.json",
    )
    for pattern in patterns:
        for path in output_dir.glob(pattern):
            suffix = path.stem.rsplit("_v", 1)[-1]
            if suffix.isdigit():
                versions.append(int(suffix))
    return max(versions) + 1 if versions else 1


def make_self_test_result() -> dict[str, Any]:
    visual_output = {
        "rule_results": [
            {
                "rule_id": rule_id,
                "status": RISK if rule_id == (VISUAL_RULE_IDS[0] if VISUAL_RULE_IDS else "") else PASS,
                "confidence": 0.94,
                "reason": "离线自检模拟画面结论。",
                "evidence": [
                    {
                        "start": "00:00:10",
                        "end": "00:00:12",
                        "detail": "离线自检模拟画面证据。",
                    }
                ],
            }
            for rule_id in VISUAL_RULE_IDS
        ]
    }
    audio_output = {
        "rule_results": [
            {
                "rule_id": rule_id,
                "status": PASS,
                "confidence": 0.93,
                "reason": "离线自检模拟音频结论。",
                "evidence": [
                    {
                        "start": "00:01:00",
                        "end": "00:01:05",
                        "detail": "离线自检模拟语音证据。",
                    }
                ],
            }
            for rule_id in AUDIO_RULE_IDS
        ],
        "script_checks": [
            {
                "item": "通用话术事项",
                "status": "matched",
                "heard_text": "已完成模拟话术事项。",
                "start": "00:00:01",
                "end": "00:00:08",
                "reason": "模拟匹配。",
            },
        ],
        "speaker_labels": [
            {
                "segment_id": "SEG-001",
                "speaker": "staff",
            },
            {
                "segment_id": "SEG-002",
                "speaker": "customer",
            },
        ],
    }
    asr_sentences = [
        {
            "segment_id": "SEG-001",
            "start": "00:00:01",
            "end": "00:00:02",
            "text": "请确认已知晓本次事项。",
        },
        {
            "segment_id": "SEG-002",
            "start": "00:00:03",
            "end": "00:00:05",
            "text": "已知晓并确认。",
        },
    ]
    metrics = {
        "duration_seconds": 120.0,
        "sample_rate": 16000,
        "channels": 1,
        "sample_width_bytes": 2,
        "rms_normalized": 0.08,
        "peak_normalized": 0.7,
        "silence_ratio": 0.25,
        "clipping_ratio": 0.0,
    }
    visual_source = normalize_rule_results(
        visual_output,
        expected_rule_ids=VISUAL_RULE_IDS,
        source="视频画面",
        confidence_threshold=0.7,
    )
    audio_source = normalize_rule_results(
        audio_output,
        expected_rule_ids=AUDIO_RULE_IDS,
        source="音频转写",
        confidence_threshold=0.7,
    )
    script_checks = normalize_script_checks(audio_output["script_checks"])
    dialogue, transcript_integrity = build_dialogue_from_asr(
        asr_sentences, audio_output["speaker_labels"]
    )
    aggregate = combine_audit_results(
        visual_source=visual_source,
        audio_source=audio_source,
        script_checks=script_checks,
        audio_metrics=metrics,
    )
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    return {
        "video": {"file_name": "self-test.mp4", "sha256": "offline", "size_bytes": 0},
            "scenario": {"id": next(iter(SCENARIOS)), "name": SCENARIOS[next(iter(SCENARIOS))]["name"]},
        "models": {
            "visual": DEFAULT_VISUAL_MODEL,
            "asr": DEFAULT_ASR_MODEL,
            "audio_rule_analysis": DEFAULT_VISUAL_MODEL,
        },
        "parameters": {
            "fps": 2.0,
            "confidence_threshold": 0.7,
            "processing_mode": "完整视频统一审核；音频内部传输后恢复连续时间轴并合并",
        },
        "started_at": now,
        "finished_at": now,
        "overall_status": overall_status(aggregate),
        "aggregate": aggregate,
        "visual": {
            "model": DEFAULT_VISUAL_MODEL,
            "rules": visual_source["rules"],
            "error": "",
        },
        "audio": {
            "asr_model": DEFAULT_ASR_MODEL,
            "analysis_model": DEFAULT_VISUAL_MODEL,
            "rules": audio_source["rules"],
            "error": "",
            "metrics": metrics,
            "script_checks": script_checks,
            "asr_sentences": sanitize_asr_sentences(asr_sentences),
            "dialogue": dialogue,
            "transcript_integrity": transcript_integrity,
        },
    }


def self_test(output_dir: Path) -> None:
    try:
        import dashscope
    except ImportError as exc:
        raise AuditError("缺少 dashscope，请先安装 requirements.txt。") from exc

    original_call = dashscope.MultiModalConversation.call
    captured_call: dict[str, Any] = {}

    def fake_streaming_call(**kwargs: Any) -> Any:
        captured_call.update(kwargs)
        return iter(
            [
                {
                    "status_code": 200,
                    "output": {
                        "choices": [
                            {"message": {"content": [{"text": '{"probe":'}]}}
                        ]
                    },
                },
                {
                    "status_code": 200,
                    "output": {
                        "choices": [
                            {"message": {"content": [{"text": "true}"}]}}
                        ]
                    },
                },
            ]
        )

    dashscope.MultiModalConversation.call = fake_streaming_call
    try:
        probe = call_qwen(
            prompt="请输出JSON对象。",
            api_key="offline-key",
            model=DEFAULT_VISUAL_MODEL,
            api_base=DEFAULT_API_BASE,
        )
    finally:
        dashscope.MultiModalConversation.call = original_call

    assert probe == {"probe": True}
    assert captured_call.get("stream") is True
    assert captured_call.get("incremental_output") is True
    assert captured_call.get("response_format") == {"type": "json_object"}
    visual_prompt = build_visual_prompt()
    assert "本次视觉规则" in visual_prompt
    assert all(rule_id in visual_prompt for rule_id in VISUAL_RULE_IDS)
    assert "未填充变量" not in visual_prompt

    repaired_time, changed = normalize_timestamp_for_duration("00:20:00", 274.11)
    assert changed and repaired_time == "00:00:20"
    repaired_time, changed = normalize_timestamp_for_duration("04:34:00", 274.11)
    assert changed and repaired_time == "00:04:34"
    repaired_time, changed = normalize_timestamp_for_duration("00:13", 334.482)
    assert changed and repaired_time == "00:00:13"
    repaired_time, changed = normalize_timestamp_for_duration("05:34", 334.482)
    assert changed and repaired_time == "00:05:34"
    assert timestamp_seconds("00:13") == 13
    assert timestamp_seconds("05:34") == 334
    repaired_reason = normalize_timestamps_in_text(
        "人像面00:13-00:28，国徽面00:35-00:39。", 334.482
    )
    assert repaired_reason == (
        "人像面00:00:13-00:00:28，国徽面00:00:35-00:00:39。"
    )
    unchanged_clock, changed = normalize_timestamp_for_duration("13:04:00", 274.11)
    assert not changed and unchanged_clock == "13:04:00"

    incremental = decumulate_asr_sentences(
        [
            {"begin_ms": 0, "end_ms": 30000, "text": "甲", "words": []},
            {"begin_ms": 0, "end_ms": 60000, "text": "甲乙", "words": []},
            {"begin_ms": 0, "end_ms": 90000, "text": "甲乙丙", "words": []},
        ]
    )
    assert [item["text"] for item in incremental] == ["甲", "乙", "丙"]
    assert [item["begin_ms"] for item in incremental] == [0, 30000, 60000]

    result = make_self_test_result()
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "self-test.json"
    excel_path = output_dir / "self-test.xlsx"
    save_json(result, json_path)
    save_excel([result], excel_path, Path("标准话术.pdf"))

    assert len(result["aggregate"]) == len(RULES)
    aggregate_by_id = {item["rule_id"]: item for item in result["aggregate"]}
    assert set(aggregate_by_id) == {rule.rule_id for rule in RULES}
    integrity = result["audio"]["transcript_integrity"]
    assert integrity["asr_segment_count"] == 2
    assert integrity["dialogue_segment_count"] == 2
    assert integrity["text_coverage"] == 1.0
    assert integrity["speaker_label_coverage"] == 1.0
    assert integrity["missing_segment_ids"] == []
    assert [item["segment_id"] for item in result["audio"]["dialogue"]] == [
        "SEG-001",
        "SEG-002",
    ]
    incomplete_dialogue, incomplete_integrity = build_dialogue_from_asr(
        result["audio"]["asr_sentences"],
        [
            {"segment_id": "SEG-001", "speaker": "staff"},
            {"segment_id": "SEG-001", "speaker": "customer"},
        ],
    )
    assert len(incomplete_dialogue) == 2
    assert incomplete_integrity["text_coverage"] == 1.0
    assert incomplete_integrity["speaker_label_coverage"] == 0.5
    assert incomplete_integrity["missing_segment_ids"] == ["SEG-002"]
    assert incomplete_integrity["duplicate_segment_ids"] == ["SEG-001"]
    assert json_path.is_file() and excel_path.is_file()
    from openpyxl import load_workbook

    test_workbook = load_workbook(excel_path, read_only=True)
    assert "原始ASR转写" in test_workbook.sheetnames
    assert "对话转写" in test_workbook.sheetnames
    assert test_workbook["原始ASR转写"]["B2"].value == "SEG-001"
    assert test_workbook["对话转写"]["B2"].value == "SEG-001"
    assert test_workbook["对话转写"]["C2"].value == "staff"
    test_workbook.close()
    print(f"离线自检通过：{json_path}")
    print(f"离线自检通过：{excel_path}")


def prepare_only(
    videos: list[Path],
    *,
    output_dir: Path,
    audio_chunk_seconds: int,
    keep_work_files: bool,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for video in videos:
        run_dir = WORK_DIR / f"prepare-{video.stem}-{uuid.uuid4().hex[:8]}"
        run_dir.mkdir(parents=True, exist_ok=False)
        try:
            prepared = prepare_audio(video, run_dir, audio_chunk_seconds)
            report = {
                "video": {
                    "file_name": video.name,
                    "sha256": sha256_file(video),
                    "size_bytes": video.stat().st_size,
                },
                "audio_metrics": prepared["metrics"],
                "audio_chunk_seconds": audio_chunk_seconds,
                "chunks": [
                    {
                        "index": chunk["index"],
                        "offset": ms_to_timestamp(chunk["offset_ms"]),
                        "duration_seconds": chunk["duration_seconds"],
                        "size_bytes": chunk["size_bytes"],
                    }
                    for chunk in prepared["chunks"]
                ],
            }
            if keep_work_files:
                report["work_files"] = {"directory": str(run_dir.resolve())}
            output_path = output_dir / f"{video.stem}_媒体准备检查.json"
            save_json(report, output_path)
            print(
                f"媒体准备检查通过：{video.name}，"
                f"音频{prepared['metrics']['duration_seconds']:.1f}秒，"
                f"共{len(prepared['chunks'])}段。"
            )
            print(f"已保存：{output_path}")
        finally:
            if not keep_work_files:
                shutil.rmtree(run_dir, ignore_errors=True)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="使用视觉模型与语音转写模型进行配置驱动的音频视觉双轨审核。"
    )
    parser.add_argument("videos", nargs="*", type=Path, help="一个或多个本地视频路径")
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help="审核配置 JSON 路径，默认 rules/audit_config.json",
    )
    parser.add_argument(
        "--script",
        type=Path,
        default=None,
        help="可选：覆盖场景配置中的话术 .md/.txt 文件",
    )
    parser.add_argument(
        "--scenario",
        default=None,
        help="场景 ID；默认使用 settings.json 的 default_scenario",
    )
    parser.add_argument(
        "--visual-model",
        default=DEFAULT_VISUAL_MODEL,
        help=f"视频画面和转写分析模型，默认{DEFAULT_VISUAL_MODEL}",
    )
    parser.add_argument(
        "--asr-model",
        default=DEFAULT_ASR_MODEL,
        help=f"音频转写模型，默认{DEFAULT_ASR_MODEL}",
    )
    parser.add_argument("--fps", type=float, default=2.0, help="视频抽帧率，默认每秒2帧")
    parser.add_argument(
        "--audio-chunk-seconds",
        type=int,
        default=180,
        help="音频内部传输块秒数，默认180秒；审核始终基于合并后的完整时间轴",
    )
    parser.add_argument(
        "--confidence-threshold",
        type=float,
        default=0.70,
        help="低于该置信度的通过结论转人工复核",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="结果目录，默认当前目录下的results",
    )
    parser.add_argument(
        "--keep-work-files",
        action="store_true",
        help="保留提取出的WAV和分段音频，便于排查",
    )
    parser.add_argument(
        "--prepare-only",
        action="store_true",
        help="只测试真实视频的音频提取和切分，不调用模型",
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="不读取视频、不调用API，验证规则合并及JSON/Excel输出",
    )
    parser.add_argument(
        "--validate-config",
        action="store_true",
        help="只检查设置、规则、场景和提示词文件，不读取视频或调用模型",
    )
    return parser


def validate_args(args: argparse.Namespace) -> None:
    if not 0.1 <= args.fps <= 10:
        raise AuditError("--fps必须在0.1到10之间。")
    if not 30 <= args.audio_chunk_seconds <= 240:
        raise AuditError("--audio-chunk-seconds必须在30到240秒之间。")
    if not 0 <= args.confidence_threshold <= 1:
        raise AuditError("--confidence-threshold必须在0到1之间。")
    if args.scenario is None:
        args.scenario = DEFAULT_SCENARIO or next(iter(SCENARIOS), None)
    if args.scenario not in SCENARIOS:
        raise AuditError("未知场景。可用场景：" + ", ".join(sorted(SCENARIOS)))
    if args.self_test or args.validate_config:
        return
    if not args.videos:
        raise AuditError("请至少提供一个视频路径，或使用--self-test。")
    for video in args.videos:
        if not video.is_file():
            raise AuditError(f"找不到视频：{video}")
        max_size_mb = int(
            ACTIVE_AUDIT_CONFIG.get("processing", {}).get(
                "max_video_size_mb", 100
            )
        )
        if video.stat().st_size > max_size_mb * 1024 * 1024:
            raise AuditError(
                f"视频{video.name}超过配置的 {max_size_mb} MB 限制，请先压缩或调整设置。"
            )
    if args.script is not None and not args.script.is_file():
        raise AuditError(f"找不到标准话术文件：{args.script}")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        load_dotenv(PROJECT_DIR / ".env")
        if (
            ACTIVE_AUDIT_CONFIG
            and Path(str(ACTIVE_AUDIT_CONFIG.get("path", ""))) == args.config.resolve()
        ):
            config = ACTIVE_AUDIT_CONFIG
        else:
            config = configure_audit_config(args.config)
        validate_args(args)
        if args.validate_config:
            for prompt_name in ("visual", "audio", "routing"):
                relative = config.get("prompts", {}).get(prompt_name)
                if not relative:
                    if prompt_name == "routing" and config.get("mode") == "advanced":
                        continue
                    raise AuditError(f"审核配置缺少 prompts.{prompt_name}。")
                prompt_path = Path(config["path"]).parent / str(relative)
                if not prompt_path.is_file():
                    raise AuditError(f"找不到提示词模板：{prompt_path}")
            print("配置检查通过：设置、规则、场景和提示词文件均可读取。")
            return 0
        if args.self_test:
            self_test(args.output_dir)
            return 0
        if args.prepare_only:
            prepare_only(
                [video.resolve() for video in args.videos],
                output_dir=args.output_dir,
                audio_chunk_seconds=args.audio_chunk_seconds,
                keep_work_files=args.keep_work_files,
            )
            return 0

        api_key = require_api_key()
        api_base = normalize_api_base(
            os.getenv(
                "MODEL_API_BASE",
                os.getenv("DASHSCOPE_BASE_HTTP_API_URL", DEFAULT_API_BASE),
            )
        )
        scenario_config = SCENARIOS[args.scenario]
        script_path = args.script or (Path(config["path"]).parent / scenario_config["script"])
        script_text = (
            extract_script_text(script_path, args.scenario)
            if script_comparison_enabled()
            else ""
        )
        args.output_dir.mkdir(parents=True, exist_ok=True)
        excel_version = next_version(args.output_dir, args.videos[0].stem)

        results: list[dict[str, Any]] = []
        for video in args.videos:
            result = audit_video(
                video.resolve(),
                script_text=script_text,
                scenario=args.scenario,
                api_key=api_key,
                api_base=api_base,
                visual_model=args.visual_model,
                asr_model=args.asr_model,
                fps=args.fps,
                audio_chunk_seconds=args.audio_chunk_seconds,
                confidence_threshold=args.confidence_threshold,
                keep_work_files=args.keep_work_files,
            )
            results.append(result)
            json_version = next_version(args.output_dir, video.stem)
            json_path = (
                args.output_dir
                / f"{video.stem}_分离审核结果_v{json_version}.json"
            )
            save_json(result, json_path)
            print(f"已保存：{json_path}")

        excel_path = (
            args.output_dir
            / f"视频审核汇总_{args.videos[0].stem}_v{excel_version}.xlsx"
        )
        save_excel(results, excel_path, script_path if script_comparison_enabled() else Path("未启用话术比对"))
        print(f"审核完成，汇总文件：{excel_path}")
        return 0
    except AuditError as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 2
    except KeyboardInterrupt:
        print("用户已中止。", file=sys.stderr)
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
